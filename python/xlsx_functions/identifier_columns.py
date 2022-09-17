import os
import re

from openpyxl import load_workbook

_ARTICLES_RECEIVER = {"al", "alla", "a", "ad", "per"}
_ARTICLES_AUTHOR = {
    "del",
    "della",
    "di",
    "sottoscritta",
    "sottoscritto",
    "da",
    "dal",
    "dalla",
    "firma",
    "firmato",
    "firmata",
}
_ARTICLES_OTHERS = {
    "sul",
    "su",
    # These are very specific words of which we know they concern 'letter subjects'
    "defunto",
    "marito",
    "Maestà",
    "fu",
    "Signore",
    "Eccellenza",
    "ricercata",
}


# pylint: disable-next=too-many-branches, too-many-locals, too-many-nested-blocks
def add_identifier_columns(
    directory_name: str, file_name: str, surnames: dict[str, list[str]]
) -> None:
    """Create and write a .xlsx file with identifier columns."""
    workbook = load_workbook(f"{directory_name}/{file_name}")
    sheet = workbook[workbook.sheetnames[0]]

    # pylint: disable-next=too-many-nested-blocks
    for index, row in enumerate(sheet.iter_rows(), start=1):
        # If there are any identifiers in the columns, assume title has been checked
        if row[6].value or row[7].value or row[8].value:
            continue

        authors, receivers, others = set(), set(), set()
        if doc_title := row[1].value:
            # Find identifiers that are not preceded by an article/word
            if len(re.findall(r"(\w+) (\$\w+)", doc_title)) != doc_title.count("$"):
                print(f"TODO: {row[0].value}:", doc_title)
                print()

            # Try to determine identifier role based on preceding word
            for article, identifier in re.findall(r"(\w+) (\$\w+)", doc_title):
                if article in _ARTICLES_AUTHOR:
                    authors.add(identifier)
                elif article in _ARTICLES_RECEIVER:
                    receivers.add(identifier)
                elif article in _ARTICLES_OTHERS:
                    others.add(identifier)
                else:
                    print(f"TODO: {row[0].value}:", doc_title)
                    print()

            # Try to find people within double-quoted titles
            if "“" in doc_title and "Giornale" not in doc_title:
                for sub_title in re.findall("“.+?”", doc_title):
                    found_surname = False
                    for surname in surnames:
                        if surname in sub_title:
                            found_surname = True
                            if len(surnames[surname]) == 1:
                                others.add(surnames[surname][0])
                                print("CHECK:", surnames[surname])
                            else:
                                print("DECIDE:", surnames[surname])
                    if found_surname:
                        print(f"{row[0].value}:", doc_title)
                        print()
                    else:
                        print(f"NONE: {row[0].value}:", doc_title)
                        print()

        # Write new cells
        sheet.cell(row=index, column=7).value = "; ".join(authors)
        sheet.cell(row=index, column=8).value = "; ".join(receivers)
        sheet.cell(row=index, column=9).value = "; ".join(others)

    new_directory = directory_name.replace("inputs", "outputs").replace(
        "VolumesExcel/", "VolumesExcelSanitized/"
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
