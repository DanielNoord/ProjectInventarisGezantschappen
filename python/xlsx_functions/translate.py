import os
import re
from re import Pattern
from typing import Literal

from openpyxl import load_workbook
from openpyxl.styles import Font
from typing_utils import Database


def translate_xlsx(
    directory_name: str,
    file_name: str,
    localization: Literal["nl_NL", "en_GB"],
    database: Database,
    used_translations: set[Pattern[str]],
) -> None:
    """Translate .xlsx file.

    Args:
        directory_name: Directory of .xlsx file
        file_name: Name of .xlsx file
        localization: Localization abbreviation
        database: Database with all translation data
        used_translations: Used translations
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        # Translate title
        if row[1].value is not None:
            line = row[1].value

            # Check if line matches one of the patterns that are automatically translated
            for pattern, trans in database.document_titles.items():
                if pattern.match(line):
                    try:
                        line = re.sub(pattern, trans[localization], line)
                    except re.error as error:
                        raise re.error(
                            f"At {pattern} found the following error: {error}"
                        ) from error
                    used_translations.add(pattern)
                    break
            else:
                row[1].font = Font(color="00008b")
                # Print out unrecognized lines
                print(f"Did not find translation for:\n{row[0].value}: {line}")
            row[1].value = line

        # Translate title
        if row[5].value is not None:
            if (place := row[5].value) in database.placenames.keys():
                place = database.placenames[place][localization]
            else:
                row[5].font = Font(color="00008b")
                print(f"Did not find placename: {place}")
            row[5].value = place

    new_directory = (
        directory_name.replace("inputs", "outputs")
        .replace("VolumesExcelSanitized/", "VolumesExcelTranslated/")
        .replace("it_IT", f"{localization}")
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name.replace('it_IT', localization)}")
    print(f"File written to {new_directory}/{file_name.replace('it_IT', localization)}")
