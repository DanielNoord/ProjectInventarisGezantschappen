import os

from openpyxl import load_workbook
from xlsx_functions.helper_functions import compare_rows


# pylint: disable-next=too-many-branches, too-many-locals, too-many-nested-blocks
def add_grouped_scans_column(directory_name: str, file_name: str) -> None:
    """Create and write a .xlsx file with identifier columns."""
    workbook = load_workbook(f"{directory_name}/{file_name}")
    sheet = workbook[workbook.sheetnames[0]]

    sheet.insert_cols(2)

    last_row = None
    # pylint: disable-next=too-many-nested-blocks
    for index, row in enumerate(sheet.iter_rows(), start=1):
        if last_row is not None and compare_rows(row, last_row, start_index=2):
            sheet.cell(row=index, column=2).value = last_row[0].value
        else:
            last_row = row
            sheet.cell(row=index, column=2).value = row[0].value

    new_directory = directory_name.replace("inputs", "outputs").replace(
        "VolumesExcel/", "VolumesExcelSanitized/"
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
