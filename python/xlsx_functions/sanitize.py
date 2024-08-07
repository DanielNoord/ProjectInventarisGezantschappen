import os

from openpyxl import load_workbook


def sanitize_xlsx(directory_name: str, file_name: str) -> None:
    """Sanitize an .xlsx file.

    Args:
        directory_name: Directory of .xlsx file
        file_name: Name of .xlsx file
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")

    for row in workbook[workbook.sheetnames[0]].iter_rows():
        for index, cell in enumerate(row[:6]):
            if cell.value is not None:
                # Clean up string
                line = cell.value
                if isinstance(line, str):
                    if line.isspace():
                        line = ""
                    elif line:
                        # Remove final characters
                        while line[-1] in {" ", ",", ".", ";"}:
                            line = line[:-1]
                        # Remove spaces in first place
                        while line[0] in {" "}:
                            line = line[1:]
                        if index:
                            if len(line) > 1:
                                line = line[0].upper() + line[1:]
                            elif len(line) == 1:
                                line = line.upper()
                        line = line.replace("( ", "(").replace(" )", ")").replace("  ", " ")
                row[index].value = line

    new_directory = directory_name.replace("inputs", "outputs").replace(
        "VolumesExcel", "VolumesExcelSanitized"
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
    print(f"File written to {new_directory}/{file_name}")
