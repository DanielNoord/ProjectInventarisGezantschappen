import os
import re
from typing import Literal

from data_parsing import name_string
from date_functions import create_document_date
from openpyxl import load_workbook
from typing_utils import Database, IndividualsDictCleaned


def fill_in_xlsx(
    directory_name: str,
    file_name: str,
    individuals: IndividualsDictCleaned,
    database: Database,
    localization: Literal["it_IT", "nl_NL", "en_GB"],
) -> None:
    """Fills an individual volume .xlsx file."""
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        if row[1].value is not None:
            date = create_document_date(row, file_name)
            line = re.split(r"( |\.|,|\(|\))", row[1].value)
            for index, word in enumerate(line):
                if word.startswith("$"):
                    try:
                        line[index] = name_string(
                            individuals[word], date, database, localization
                        )
                    except KeyError as error:
                        raise KeyError(
                            f"Incorrect identifier {error} in {file_name} on row {row[0].row}"
                        ) from error
            row[1].value = "".join(line)

    new_directory = (
        directory_name.replace("VolumesExcelSanitized/", "VolumesExcelFinal/")
        .replace("VolumesExcelTranslated/", "VolumesExcelFinal/")
        .replace("inputs", "outputs")
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
    print(f"File written to {new_directory}/{file_name}")
