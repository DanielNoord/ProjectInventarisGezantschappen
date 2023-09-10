import re
from collections.abc import Iterator
from pathlib import Path
from typing import NewType

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from xlsx_functions.helper_functions import compare_rows

VolumeIdentifier = NewType("VolumeIdentifier", str)
FileIdentifier = NewType("FileIdentifier", str)


class FileScanMapper:
    def __init__(self, file_name: Path) -> None:
        self._file_name = file_name

    def run(self) -> dict[FileIdentifier, list[str]]:
        """Get a mapping of all the files and scans in the file."""
        files: dict[FileIdentifier, list[str]] = {}
        current_file: tuple[Cell, ...] | None = None

        for prevr, row, nextr in self.get_rows():
            # Sometimes we added an empty row to separate series. We skip those, but
            # make sure to reset the current file as we are sure a new file is started
            if (file := row[0].value) is None:
                current_file = None
                continue

            assert isinstance(file, str)

            # Handle new volumes
            if file.endswith("_title"):
                # Reset the current file, as a file can never span multiple volumes
                current_file = None
                vol = VolumeIdentifier(file.replace("_title", ""))
                if not vol.count("_"):
                    # Add the dorso and p... (front) scan
                    files[FileIdentifier(vol + "_d")] = [vol + "_d"]
                    files[FileIdentifier(vol + "_p")] = [vol + "_p"]
                continue

            # The current row is now certainly a file
            file = FileIdentifier(file)

            # If there is no difference with the current file, append to it
            if compare_rows(current_file, row):
                for to_add in self.get_scans_to_add(
                    str(prevr[0].value) if prevr is not None else prevr,
                    file,
                    str(nextr[0].value) if nextr is not None else nextr,
                ):
                    files[FileIdentifier(current_file[0].value)].append(to_add)
            else:
                # Create a new file and append to it
                current_file = row
                files[file] = []
                for to_add in self.get_scans_to_add(
                    str(prevr[0].value) if prevr is not None else prevr,
                    file,
                    str(nextr[0].value) if nextr is not None else nextr,
                ):
                    files[file].append(to_add)

        return files

    @staticmethod
    def get_scans_to_add(
        prevr: str | None, row: str, nextr: str | None
    ) -> Iterator[str]:
        """Get the file names of the scans to add depending on the preceding and next row."""
        # If the row is a specific Verso row and not part of a sequence following u we only
        # add the verso scan.
        if row.endswith("v") and prevr and not prevr.endswith("u"):
            yield row
        else:
            # We always add the recto scan. If the next row is not a verso-only file we also
            # add the verso scan.
            yield row + "r"
            if not (nextr and nextr.endswith("v") and not row.endswith("u")):
                yield row + "v"

    def get_rows(
        self,
    ) -> Iterator[
        tuple[tuple[Cell, ...] | None, tuple[Cell, ...], tuple[Cell, ...] | None]
    ]:
        """Open a workbook and return all rows together with their preceding and next row."""
        workbook = load_workbook(self._file_name)
        sheet = workbook[workbook.sheetnames[0]]
        # Little sanity check to make sure the first row is a title row
        if not re.match("(.*)_title", sheet["A"][0].value):
            raise ValueError(
                f"Can't determine the volume/ms number of {sheet['A'][0].value}. "
                "Does it have the correct format?"
            )

        rows = list(sheet.iter_rows())
        max_rows = len(rows) - 1
        for i, row in enumerate(rows):
            if i == 0:
                yield None, row, rows[i + 1]
            elif i == max_rows:
                yield rows[i - 1], row, None
            else:
                yield rows[i - 1], row, rows[i + 1]

        # Close the workbook again
        workbook.close()
