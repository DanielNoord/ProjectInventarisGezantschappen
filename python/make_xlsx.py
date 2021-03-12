import json
import os

from openpyxl import load_workbook

from functions.create_name_string import name_string
from functions.translate import initialize_translation_database


def translate_xlsx(directory_name, file_name, individuals, translations, localization):
    """Translates and writes an individual volume .xlsx file

    Args:
        directory_name (str): Name of input directory
        file_name (str): Name of input file
        individuals (dict): Dictionary of all individuals based on inputs/Individuals.json
        translations (list): Dictionaries of function and title translations
        localization (str): Localization abbreviation ("nl_NL", "it_IT", "en_GB")
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        if row[1].value is not None:
            date = [row[2].value, row[3].value, row[4].value]
            line = row[1].value.split()
            for index, word in enumerate(line):
                if word.startswith("$"):
                    line[index] = name_string(
                        individuals[word], date, translations, localization
                    )
            row[1].value = " ".join(line)

    os.makedirs(
        os.path.join(os.getcwd(), directory_name.replace("inputs", "outputs")),
        exist_ok=True,
    )
    workbook.save(f"{directory_name.replace('inputs', 'outputs')}/{file_name}")
    print(f"File written to {directory_name.replace('inputs', 'outputs')}/{file_name}")


def create_xlsx(directory_name, localization):
    """Creates the .xlsx files of all volumes in the input directory

    Args:
        directory_name (str): Name of the directory with the input .xlsx files
        localization (str): Localization abbreviation ("nl_NL", "it_IT", "en_GB")
    """
    with open("inputs/Individuals.json") as file:
        individuals_data = json.load(file)
    del individuals_data["$schema"]
    translation_data = initialize_translation_database()

    directory = os.fsencode(directory_name)
    for file in os.listdir(directory):
        if not str(file).count("~$") and str(file).startswith("b'Paesi"):
            filename = os.fsdecode(file)
            translate_xlsx(
                directory_name,
                filename,
                individuals_data,
                translation_data,
                localization,
            )


if __name__ == "__main__":
    create_xlsx("inputs/VolumesExcel/en_GB", "en_GB")
    create_xlsx("inputs/VolumesExcel/it_IT", "it_IT")
    create_xlsx("inputs/VolumesExcel/nl_NL", "nl_NL")
