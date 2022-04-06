#!/usr/bin/env python3

import os
import re
from typing import cast

import openpyxl
from lxml import etree
from openpyxl import load_workbook

from xlsx_functions import compare_rows, parse_file, parse_series
from xlsx_make import create_sanitized_xlsx
from xml_functions import XMLWriter, add_dao, fix_daoset


# pylint: disable-next=too-few-public-methods
class EADMaker(XMLWriter):
    """Class which can a EAD compliant .xml file."""

    def __init__(self, input_dir: str, sanitize: bool = True) -> None:
        super().__init__()

        # Input and output locations attributes
        self.sanitized_dir = input_dir.replace("inputs", "outputs").replace(
            "VolumesExcel", "VolumesExcelSanitized"
        )
        """Directory with input .xlsx files."""

        self.xml_file = "outputs/Legation_Archive.xml"
        """Filename of the final .xml file."""

        # Sanitize the input .xlsx files
        if sanitize:
            create_sanitized_xlsx(input_dir)

        self.series: dict[str, dict[str, etree._Element]] = {}
        """All series and their etree elements within the archive."""

    def create_ead(self, *, print_identifier_count: bool = False) -> None:
        """Start the process of the EAD creation."""
        self._create_xml_file()
        self._xml_check()

        if print_identifier_count:
            count = sorted(
                [(v, k) for k, v in self.identifier_counter.items()],
                key=lambda tup: -tup[0],
            )
            for i in count:
                print(i)

    def _write_xml_file(self, root: etree._Element) -> None:
        """Write a .xml file from a an etree Element."""
        tree = etree.ElementTree(root)

        # Check if outputs directory exists and then write file
        os.makedirs(os.path.join(os.getcwd(), r"outputs"), exist_ok=True)
        tree.write(  # type: ignore[call-arg] # Stub doesn't recognize doctype parameter is valid
            self.xml_file,
            pretty_print=True,
            xml_declaration=True,
            encoding="UTF-8",
            doctype="""<!DOCTYPE ead PUBLIC "+// http://ead3.archivists.org/schema/ //DTD ead3 (Encoded Archival Description (EAD) Version 3)//EN" "ead3.dtd">""",  # pylint: disable=line-too-long
        )

        print(f"Printed file to {self.xml_file}")

    def _xml_check(self) -> None:
        """Perform .xml check with xmllint."""
        os.system(
            # pylint: disable-next=line-too-long
            f"xmllint --noout --dtdvalid outputs/ead3.dtd {self.xml_file} 2> {self.log_xml_errors}"
        )
        with open(self.log_xml_errors, encoding="utf-8") as file:
            assert not file.read()
        print("XML-DTD check complete!")

    def _create_xml_file(self) -> None:
        """Create and write an xml file based on a directory of volumes."""
        print("Starting to create XML file!")

        root, archdesc_dsc = self.basic_xml_file()

        # Iterate files sorted by volume number
        files = [i for i in os.listdir(self.sanitized_dir) if i.startswith("Paesi")]
        for file in sorted(
            files,
            key=lambda name: int(
                name.replace("Paesi Bassi VOLUME ", "").replace("_it_IT.xlsx", "")
            ),
        ):
            self._create_xml_series(f"{self.sanitized_dir}/{file}", archdesc_dsc)

        fix_daoset(root)

        self._write_xml_file(root)

        print("Writing XML complete!")

        # Printing some descriptive stats
        print("Found the following dossiers:")
        for i in self.series.values():
            print(", ".join(i))
        print("Found the following unused translations:")
        print(
            "\n".join(
                i.pattern
                for i in self.database.document_titles.keys()
                if i not in self.used_translations
            )
        )

    def _create_xml_series(self, filename: str, archdesc: etree._Element) -> None:
        """Add a volume to an 'archdesc' element."""
        workbook = load_workbook(filename)
        first_sheet = workbook[workbook.sheetnames[0]]
        sub_levels: dict[str, etree._Element] = {}

        for index, cell in enumerate(first_sheet["A"]):
            if match := re.match("(.*)_title", cell.value or ""):
                series_data = parse_series(first_sheet[index + 1])
                if series_data.level == 1:
                    sub_levels[match.groups()[0]] = self.series_entry(
                        archdesc, series_data
                    )
                else:
                    if not (parent_level := re.match(r"(.*)_.*?", match.groups()[0])):
                        raise ValueError(
                            f"Can't determine the series parent of {cell.value}."
                            "Does it have the correct format?"
                        )
                    sub_levels[match.groups()[0]] = self.series_entry(
                        sub_levels[parent_level.groups()[0]], series_data
                    )

        self._create_xml_individual_files(first_sheet, sub_levels)

        if not (volume_number := re.match("(.*)_title", first_sheet["A"][0].value)):
            raise ValueError(
                f"Can't determine the volume/ms number of {first_sheet['A'][0].value}."
                "Does it have the correct format?"
            )
        print(
            f"Finished writing volume {volume_number.groups()[0]} "
            f"- Volume {int(volume_number.groups()[0][2:]) - 275}"
        )

        self.series.update({str(volume_number.groups()[0]): sub_levels})

    def _create_xml_individual_files(  # pylint: disable=too-many-branches
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        sub_series: dict[str, etree._Element],
    ) -> None:
        """Based on a sheet creates .xml element entries for every file found."""
        prev_file, prev_file_did, prev_series = None, None, None

        for file in sheet.iter_rows():
            similar = False
            individual_verso = False

            # Skip empty lines or series title lines
            if file[0].value is None or file[0].value.endswith("_title"):
                continue

            # Error on files with a space in their "file number"
            if " " in file[0].value:
                raise ValueError(
                    f"There is a space in the file number {file[0].value}. This is not allowed!"
                )

            file_data = parse_file(file)

            if (
                prev_file is not None
                and prev_file_did is not None
                and file_data.series == prev_series
                and file_data.title != "Bianca"
            ):
                similar = compare_rows(file, prev_file)

            # If current file is a verso description, remove verso from previous daoset
            if re.match(r".+v", file_data.file_name):
                if prev_file_did is None:
                    raise ValueError(
                        # pylint: disable-next=line-too-long
                        f"{file_data.file_name} is a verso appearing before the description of {file_data.file_name[:-1]}"
                    )

                # If the previous file ends in u, the v is not 'verso', but
                # continuation of long document.
                if not prev_file_did.find("unitid").text.endswith("u"):
                    individual_verso = True

                if individual_verso:
                    for dao in prev_file_did.find("daoset"):
                        if dao.attrib["id"] == f"{file_data.file_name}.tif":
                            dao.getparent().remove(dao)

            if not similar:
                prev_file_did = self.file_entry(
                    sub_series[file_data.series], file_data, individual_verso
                )
            # Update pages/id of previous document
            else:
                # If similar means prev_file_did is defined
                prev_file_did = cast(
                    etree._Element, prev_file_did  # pylint: disable=protected-access
                )
                unitid = prev_file_did.find("unitid")
                if (
                    not isinstance(
                        unitid, etree._Element  # pylint: disable=protected-access
                    )
                    or not unitid.text
                    or not isinstance(unitid.text, str)
                ):
                    raise ValueError(
                        f"Can't find unitid in {prev_file_did}, it is empty or it isn't a string"
                    )
                if "-" in unitid.text:
                    unitid.text = (
                        unitid.text[: unitid.text.index("-") + 1] + file_data.page
                    )
                else:
                    unitid.text += f"-{file_data.page}"

                # Update daoset of previous document
                daoset = prev_file_did.find("daoset")
                if not isinstance(
                    daoset, etree._Element  # pylint: disable=protected-access
                ):
                    raise ValueError(f"Can't find daoset in {prev_file_did}")
                add_dao(daoset, file_data, individual_verso)

            prev_file = file
            prev_series = file_data.series


if __name__ == "__main__":
    eadmaker = EADMaker(
        "inputs/VolumesExcel/it_IT",
        True,
    )
    eadmaker.create_ead()
