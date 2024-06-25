import json
import os
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from xlsx_make import create_sanitized_xlsx

_HUMAN_READABLE_COLUMN_NAMES: dict[int, str] = {
    1: "Title",
    2: "Year",
    3: "Month",
    4: "Day",
    5: "Place",
    6: "Author",
    7: "Recipient",
    8: "Subject",
}


class PartialMatcher:

    def __init__(self) -> None:
        self.partial_matches: dict[int, defaultdict[str, list[str]]] = {
            8: defaultdict(list),
            7: defaultdict(list),
            6: defaultdict(list),
            5: defaultdict(list),
            4: defaultdict(list),
            3: defaultdict(list),
            2: defaultdict(list),
            1: defaultdict(list),
        }
        self.total_count = 0

    def run(self, print_to_file: bool, sanitize: bool) -> None:
        """Find all rows that are partial matches."""
        input_dir = Path("inputs") / "VolumesExcel" / "it_IT"
        if sanitize:
            create_sanitized_xlsx(str(input_dir))
        sanitized_dir = (
            str(input_dir)
            .replace("inputs", "outputs")
            .replace("VolumesExcel", "VolumesExcelSanitized")
        )
        files = [i for i in os.listdir(sanitized_dir) if i.startswith("Paesi")]

        for file in sorted(
            files,
            key=lambda name: int(
                name.replace("Paesi Bassi VOLUME ", "").replace("_it_IT.xlsx", "")
            ),
        ):
            self._find_partial_match(file, sanitized_dir)

        print(self.total_count)

        final_dict = {
            k: {i: {"count": len(j), "files": j} for i, j in v.items()}
            for k, v in self.partial_matches.items()
            if k != 8
        }

        with open("outputs/partial_matches.json", "w", encoding="utf-8") as file:
            json.dump(final_dict, file, ensure_ascii=False, indent=4)

    def _find_partial_match(self, file: str, sanitized_dir: str) -> None:
        workbook = load_workbook(Path(sanitized_dir) / file)
        first_sheet = workbook[workbook.sheetnames[0]]

        prev_row = None
        for row in first_sheet.iter_rows(values_only=True):
            self.total_count += 1

            if prev_row is None or row[0] is None:
                prev_row = row
                continue

            matching_indices: set[int] = set()
            for index, (old_value, new_value) in enumerate(zip(prev_row, row)):
                # We only care about the first 7 columns
                if index > 7:
                    break

                if old_value == new_value:
                    matching_indices.add(index)

            for i in range(index, 9):
                matching_indices.add(i)

            if matching_indices:
                self.partial_matches[len(matching_indices)][
                    ", ".join(
                        _HUMAN_READABLE_COLUMN_NAMES[i]
                        for i in sorted(matching_indices)
                    )
                ].append(str(row[0]))

            prev_row = row
