import os
import re
from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook


class FileRenamer:
    """Renames files based on arguments."""

    def __init__(self, folder: str, scan_dir: str, test: bool) -> None:
        self.filename_pattern = r"(\d+)(bis)*([rv].tif)"
        """Pattern of the file names to match against."""

        self.folder = folder + scan_dir
        """Folder to scan."""

        self.volume = scan_dir

        self.renamed = 0
        """Amount of renamed files."""

        self.test_run = test
        """Whether this is a dry run without renaming any files."""

    def rename_files(self, rename_dict: dict[Callable[[int], bool], str], sub_dir: str) -> None:
        """Rename files in a folder based on attributes."""
        for file in os.listdir(self.folder):
            if file.endswith("00"):
                continue

            sanitized_file = (
                file.replace("rbis", "bisr")
                .replace("vbis", "bisv")
                .replace("_r.tif", "r.tif")
                .replace("_v.tif", "v.tif")
                .replace("V.tif", "v.tif")
            )
            if match := re.match(
                self.volume + "_" + sub_dir + self.filename_pattern, sanitized_file
            ):
                file_num = int(match.groups()[0])
                for value_test, dossier in rename_dict.items():
                    if value_test(file_num):
                        oldname = f"{self.folder}/{file}"
                        newfile = sanitized_file.replace(sub_dir, "").replace(
                            self.volume, f"{self.volume}{dossier}"
                        )
                        newname = f"{self.folder}/{newfile}"

                        if self.test_run:
                            print("Would rename:")
                            print(oldname)
                            print(newname)
                        else:
                            os.rename(oldname, newname)
                            print(f"Renamed: {oldname} -> {newname}")
                        self.renamed += 1

                        break

        if self.test_run:
            print(f"Would rename {self.renamed} files.")
        else:
            print(f"Renamed {self.renamed} files.")


class FileRenamerExcel:
    """File renamer based on  names in Excel."""

    def __init__(self, excel_dir: str, scans_dir: str, test: bool) -> None:
        self.excel_files: list[str] = []
        """All filenames of the excel files."""

        for file in os.listdir(excel_dir):
            if file.endswith(".xlsx") and not file.startswith("~$"):
                self.excel_files.append(f"{excel_dir}/{file}")

        self.excel_files = sorted(self.excel_files)

        self.scans_dir = scans_dir
        """Location of the scans."""

        self.test_run = test
        """Whether this is a dry run without renaming any files."""

        self.renamed = 0
        """Amount of renamed files."""

    def _get_rename_pairs(self, filename: str) -> tuple[str, dict[str, str]]:
        """Get the rename pairs from a workbook and remove it from the row."""
        workbook = load_workbook(filename)
        first_sheet = workbook[workbook.sheetnames[0]]

        # Get the volume name
        folder = first_sheet["A"][0].value.split("_")[0].upper()
        assert isinstance(folder, str)
        assert folder.startswith("MS")

        rename_pairs: dict[str, str] = {}
        for row in first_sheet.iter_rows():
            # Only check long enough rows
            if not len(row) >= 10:
                continue

            # See if there is a scan name
            if not (scan_name := row[9].value):
                continue

            assert isinstance(scan_name, str)
            scan_name = scan_name.rstrip()
            if scan_name.count("r") > 1:
                if self.test_run:
                    print(f"DEBUG: Double 'r' in {scan_name}")
                else:
                    raise ValueError(f"We can't handle names with two 'r's. See {scan_name}")
            if scan_name in rename_pairs:
                if self.test_run:
                    print(f"DEBUG: Double scan name for {scan_name}")
                else:
                    raise ValueError(f"Double scan name found: {scan_name}")
            new_name = row[0].value.replace("ms", "MS")
            rename_pairs[scan_name] = f"{new_name}r"
            rename_pairs[scan_name.replace("r", "v")] = f"{new_name}v"
            row[9].value = None

        workbook.save(filename.replace("inputs", "outputs"))

        return folder, rename_pairs

    def _get_files_to_rename(self) -> dict[str, dict[str, str]]:
        """Iterate over all excel_files and check for renameable scans."""
        to_rename: dict[str, dict[str, str]] = {}
        os.makedirs(
            str(Path(self.excel_files[0]).parent).replace("inputs", "outputs"),
            exist_ok=True,
        )
        for filename in self.excel_files:
            if not self.test_run:
                print(f"Checking {filename}.")
            folder, rename_pairs = self._get_rename_pairs(filename)
            to_rename[folder] = rename_pairs
        return to_rename

    def _check_scans_exist(self, rename_pairs: dict[str, dict[str, str]]) -> None:
        """Check if scans to rename actually exist."""
        for folder, pairs in rename_pairs.items():
            if pairs:
                foldername = f"{self.scans_dir}/{folder}"
                for old in pairs:
                    oldname = f"{foldername}/{old}.tif"
                    if not os.path.exists(oldname):
                        if self.test_run:
                            print(f"DEBUG: Can't find scan {old}")
                        else:
                            raise ValueError(f"Missing scan for {oldname}")

    def _rename_scans(self, rename_pairs: dict[str, dict[str, str]]) -> None:
        """Rename scans based on rename pairs."""
        for folder, pairs in rename_pairs.items():
            if pairs:
                foldername = f"{self.scans_dir}/{folder}"
                for old, new in pairs.items():
                    oldname = f"{foldername}/{old}.tif"
                    newname = f"{foldername}/{new}.tif"
                    self.renamed += 1
                    if self.test_run:
                        print(f"Would rename: {oldname} -> {newname}")
                    else:
                        print(f"Renamed: {oldname} -> {newname}")
                        os.rename(oldname, newname)
        print(f"Renamed {self.renamed} scans.")

    def rename_files(self) -> None:
        """Full loop of renaming scan files."""
        to_rename = self._get_files_to_rename()
        self._check_scans_exist(to_rename)
        self._rename_scans(to_rename)


if __name__ == "__main__":
    # Uncomment to run file renamer in test mode.
    # renamer = FileRenamer(
    #     "/Volumes/Seagate Basic Media/VolumesLegazione/",
    #     "MS293",
    #     test=True,
    # )
    # renamer.rename_files(
    #     {
    #         lambda x: 0 <= x <= 61: "_20_1",
    #         # lambda x: 62 <= x <= 63: "_20_2",
    #         # lambda x: 64 <= x <= 78: "_20_3",
    #         # lambda x: 79 <= x <= 86: "_20_4",
    #         # lambda x: 87 <= x <= 88: "_20_5",
    #         # lambda x: 89 <= x <= 136: "_20_6",
    #         # lambda x: 137 <= x <= 194: "_20_7",
    #         # lambda x: 195 <= x <= 232: "_20_8",
    #         # lambda x: 233 <= x <= 240: "_20_8",
    #         # lambda x: 241 <= x <= 250: "_20_9",
    #         # lambda x: 251 <= x <= 260: "_20_10",
    #         # lambda x: 261 <= x <= 305: "_20_21",
    #         # lambda x: 306 <= x <= 410: "_30_4",
    #     },
    #     "20_",
    # )

    # Uncomment to run excel file renamer in test mode.
    # renamer = FileRenamerExcel(
    #     "inputs/VolumesExcel/it_IT",
    #     "/Volumes/Seagate Basic Media/VolumesLegazione",
    #     test=True,
    # )
    # renamer.rename_files()

    print()
