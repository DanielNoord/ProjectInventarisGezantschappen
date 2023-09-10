#!/usr/bin/env python3

from __future__ import annotations

import csv
import json
import os
import re
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from contentdm.file_scan_mapper import FileScanMapper
from data_parsing import control_title
from data_parsing.load_database import initialize_database_for_xml
from openpyxl import load_workbook
from typing_utils.translations_classes import Database
from xlsx_make import create_sanitized_xlsx
from xml_functions.title_elements import fill_in_name, fix_quotes


@dataclass
class FileRow:  # pylint: disable=too-many-instance-attributes
    scans: list[str]
    file_id: str
    series_id: str
    title: str
    year: int | None
    month: int | None
    day: int | None
    location: str | None
    authors: list[str]
    recipients: list[str]
    subjects: list[str]

    @classmethod
    def from_row(cls, row: list[openpyxl.cell.Cell], scans: list[str]) -> FileRow:
        # Check for empty lines, title lines or incorrect lines
        file_id = row[0].value
        if file_id is None:
            file_id = ""
        assert isinstance(file_id, str)
        if not file_id or file_id.endswith("_title") or " " in file_id:
            raise ValueError(f"Invalid file number '{file_id}'.")

        # Get series ID
        if mat := re.match(r"(.*)_(.*)", file_id):
            series_id = mat.groups()[0]
        else:
            raise ValueError(f"Can't parse series ID of: '{file_id}'")

        # Cast all rows to their expected type
        title = str(row[1].value) if row[1].value is not None else None
        year = int(row[2].value) if row[2].value is not None else None
        month = int(row[3].value) if row[3].value is not None else None
        day = int(row[4].value) if row[4].value is not None else None
        location = str(row[5].value) if row[5].value is not None else None
        authors = str(row[6].value) if row[6].value is not None else None
        recipients = str(row[7].value) if row[7].value is not None else None
        subjects = str(row[8].value) if row[8].value is not None else None

        return cls(
            scans=scans,
            file_id=file_id,
            series_id=series_id,
            title=title or "",
            year=year,
            month=month,
            day=day,
            location=location,
            authors=authors.split("; ") if authors else [],
            recipients=recipients.split("; ") if recipients else [],
            subjects=subjects.split("; ") if subjects else [],
        )

    @property
    def date(self) -> str:
        """The date in the format xxxx-xx-xx."""
        year, month, day = self.year, self.month, self.day
        if month and not year:
            month = None
        if day and not month:
            day = None
        return "-".join(str(i).zfill(2) for i in (year, month, day) if i)


# pylint: disable-next=too-few-public-methods
class ContentDMFileWriter:
    """Class which can write a tab-delimited txt file to be imported by ContentDM."""

    def __init__(self, input_dir: str, sanitize: bool = True) -> None:
        # Load the database with translations and individuals
        self.database = initialize_database_for_xml()
        """Database with all translations and individuals."""

        self.sanitized_dir = Path(
            input_dir.replace("inputs", "outputs").replace(
                "VolumesExcel", "VolumesExcelSanitized"
            )
        )
        """Directory with input .xlsx files."""

        self.csv_file = Path("outputs/Legation_Archive.txt")
        """Filename of the final .txt file."""

        # Sanitize the input .xlsx files
        if sanitize:
            create_sanitized_xlsx(input_dir)

    def run(self) -> None:
        """Create a TSV file for ContentDM."""
        output_dir = Path("outputs") / "contentdm"
        for griglie in self._get_all_griglie():
            os.makedirs(output_dir / griglie.stem, exist_ok=True)

            for file in self._get_all_files(griglie):
                file_dir = output_dir / griglie.stem / file.file_id
                os.makedirs(file_dir / "scans", exist_ok=True)
                with open(file_dir / "metadata.txt", "w", encoding="utf-8") as tsv_file:
                    row = self._get_actual_excel_row(file)
                    row["filename"] = None
                    tsv_writer = csv.DictWriter(
                        tsv_file, dialect=csv.excel_tab, fieldnames=list(row.keys())
                    )
                    tsv_writer.writeheader()
                    tsv_writer.writerow(row)
                    for scan in file.scans:
                        row["filename"] = scan
                        tsv_writer.writerow(row)

    def _get_all_griglie(self) -> Iterator[Path]:
        """Yields all file paths to the santizied griglie in sorted order"""
        files = [
            (i, i.name.replace("Paesi Bassi VOLUME", "").replace("_it_IT.xlsx", ""))
            for i in self.sanitized_dir.iterdir()
            if i.name.startswith("Paesi")
        ]
        for file, _ in sorted(files, key=lambda x: int(x[1])):
            if not 29 < int(_) < 33:
                continue
            yield file
            print(f"Finished parsing {file.name}")

    def _get_all_files(self, griglie: Path) -> Iterator[FileRow]:
        files_and_scans = FileScanMapper(griglie).run()

        workbook = load_workbook(griglie)
        sheet = workbook[workbook.sheetnames[0]]

        for index, row in enumerate(sheet.iter_rows()):
            if index == 0:
                original_name = str(row[0].value).removesuffix("_title")
                for suffix in ["_d", "_p"]:
                    name = original_name + suffix
                    row[0].value = name
                    yield FileRow.from_row(row, [name])
            if row[0].value in files_and_scans:
                yield FileRow.from_row(row, files_and_scans[row[0].value])

    def _get_actual_excel_row(
        self, file: FileRow
    ) -> dict[str, str | int | dict[str, str | int]]:
        """Turn an ImageRow into a dict with the correct column headers."""
        title_nl, title_en, title_it = self._translate_title(file, self.database)
        loc_data = self.database.placenames[file.location] if file.location else None
        authors = {k: self.database.individuals[k] for k in file.authors}
        recipients = {k: self.database.individuals[k] for k in file.recipients}
        subjects = {k: self.database.individuals[k] for k in file.subjects}

        return {
            "title_en_gb": title_en,
            "title_it_it": title_it,
            "title_nl_nl": title_nl,
            "title_identifier": file.title or "",
            "location": {
                "location_en_gb": loc_data["nl_NL"] if loc_data else "",
                "location_it_it": file.location or "",
                "location_nl_nl": loc_data["nl_NL"] if loc_data else "",
                "geonames_id": loc_data["geonames_id"] if loc_data else "",
                "longitude": loc_data["longitude"] if loc_data else "",
                "latitude": loc_data["latitude"] if loc_data else "",
            },
            "year": file.year or "",
            "month": file.month or "",
            "day": file.day or "",
            "authors": json.dumps(authors),
            "recipients": json.dumps(recipients),
            "subjects": json.dumps(subjects),
            "series_data": "I have not found a good way to serialize this.",
        }

    def _translate_title(
        self, file: FileRow, database: Database
    ) -> tuple[str, str, str]:
        """Translate a title into Dutch, English and Italian."""
        # Find document translation
        for pattern, trans in database.document_titles.items():
            if pattern.match(file.title):
                try:
                    title_en = re.sub(pattern, trans["en_GB"], file.title)
                    title_nl = re.sub(pattern, trans["nl_NL"], file.title)
                    break
                except re.error as error:
                    raise re.error(
                        f"At {pattern} found the following error: {ersror}"
                    ) from error
        else:
            raise ValueError(f"Could not find a translation for {file.title}")

        title_it = fill_in_name(file.title, database, file.date, "it_IT")
        title_en = fill_in_name(title_en, database, file.date, "en_GB")
        title_nl = fill_in_name(title_nl, database, file.date, "nl_NL")

        # Add and check italics
        if sum("_" in i for i in (title_it, title_en, title_nl)) == 1:
            raise ValueError(
                f"Only one language has italtics indication for {file.title}"
            )

        if re.search(r"\"|“|”", title_it):
            title_it = fix_quotes(title_it)
            title_en = fix_quotes(title_en)
            title_nl = fix_quotes(title_nl)

        control_title(title_it, file.title)
        control_title(title_en, file.title)
        control_title(title_nl, file.title)

        return title_nl, title_en, title_it


if __name__ == "__main__":
    eadmaker = ContentDMFileWriter("inputs/VolumesExcel_06_07_2022/it_IT", False)
    eadmaker.run()
