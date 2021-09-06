#!/usr/bin/env python3

import os
import re
from warnings import warn

import openpyxl
from lxml import etree
from openpyxl import load_workbook

from functions.xlsx_parse import parse_dossier, parse_file, parse_volume
from functions.xml_create_elements import basic_xml_file, dossier_entry, file_entry, volume_entry


def create_xml_individual_files(
    localization: str,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    dossiers: dict,
    vol_entry: etree._Element,
) -> None:
    """Based on a sheet creates .xml entries for every file found

    Args:
        localization (str): String indicating the localization of the inventory
        sheet (openpyxl.worksheet.worksheet.Worksheet): The .xlsx sheet with data
        dossiers (dict): Dictionary of dossier numbers with their corresponding .xml element
        vol_entry (etree._Element): The c01 element of the final .xml file
            To be used when there is no dossier.
    """
    for file in sheet.iter_rows():
        if file[0].value is not None and not file[0].value.endswith("_0"):
            f_page, f_title, f_place, f_date = parse_file(file)

            # Check if file belongs to a dossier
            if mat := re.search(r"ms.*?_(.*?)_.*?", file[0].value):
                file_entry(
                    dossiers[mat.groups()[0]],
                    f_page,
                    f_title,
                    f_place,
                    f_date,
                    localization,
                )
            else:
                file_entry(vol_entry, f_page, f_title, f_place, f_date, localization)


def create_xml_dossier(
    localization: str,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    v_num: str,
    c01: etree._Element,
) -> None:
    """Creates necessary dossier entries at the c02 level

    Args:
        localization (str): String indicating the localization of the inventory
        sheet (openpyxl.worksheet.worksheet.Worksheet): The .xlsx sheet with data
        v_num (str): Number of the volume
        c01 (etree._Element): The c01 element of the final .xml file
    """
    dossiers = {}

    # Find dossiers
    for cell in sheet["A"]:
        if cell.value is not None and (mat := re.search(r"ms.*?_(.*?)_.*?", cell.value)):
            if mat.groups()[0] not in dossiers.keys():
                d_pages, d_title, d_data = parse_dossier(sheet, mat.groups()[0], v_num, cell)

                # Create entry
                dossiers[mat.groups()[0]] = dossier_entry(
                    c01,
                    v_num,
                    mat.groups()[0],
                    d_pages,
                    d_title,
                    d_data,
                    localization,
                )

    # Create dossier if no dossiers were found?
    # TODO: Do we want to create a dossier in these cases?
    if dossiers == {}:
        warn(f"V{v_num} does not have any dossiers!")

    create_xml_individual_files(localization, sheet, dossiers, c01)


def create_xml_volume(localization: str, filename: str, archdesc: etree._Element) -> None:
    """Adds a volume to an 'archdesc' element

    Args:
        localization (str): String indicating the localization of the inventory
        filename (str): Name and directory of file
        archdesc (etree._Element): The archdesc element of the final .xml file
    """
    workbook = load_workbook(filename)
    first_sheet = workbook[workbook.sheetnames[0]]

    # Create volume entry at c01 level
    v_num, v_title, v_date = parse_volume(first_sheet[1])
    c01 = volume_entry(archdesc, v_num, v_title, v_date, localization)

    create_xml_dossier(localization, first_sheet, v_num, c01)

    print(f"""Finished writing volume {v_num}\n""")


def create_xml_file(localization: str, dir_name: str) -> None:
    """Creates and writes an xml file based on directory of volumes

    Args:
        localization (str): String indicating the localization of the inventory
        dir_name (str): Directory name
    """
    print("Starting to create XML file!")
    root, archdesc = basic_xml_file()

    directory = os.fsencode(f"{dir_name}{localization}")
    for file in sorted(os.listdir(directory)):
        filename = os.fsdecode(file)
        if not filename.count("~$") and filename.startswith("Paesi"):
            create_xml_volume(localization, f"{dir_name}{localization}/{filename}", archdesc)

    tree = etree.ElementTree(root)

    # Check if outputs directory exists and then write file
    os.makedirs(os.path.join(os.getcwd(), r"outputs"), exist_ok=True)
    tree.write(
        f"outputs/Inventaris_{localization}.xml",
        pretty_print=True,
        xml_declaration=True,
        encoding="UTF-8",
        doctype="""<!DOCTYPE ead SYSTEM "http://www.nationaalarchief.nl/collectie/ead/ead.dtd">""",
    )

    print(f"Wrote file to outputs/Inventaris_{localization}.xml")
    print("Writing XML complete!")


if __name__ == "__main__":
    create_xml_file("it_IT", "outputs/VolumesExcelFinal/")
