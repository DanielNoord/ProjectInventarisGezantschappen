#!/usr/bin/env python3

import os
import re
from warnings import warn

import openpyxl
from lxml import etree
from openpyxl import load_workbook

from xlsx_functions import parse_dossier, parse_file, parse_volume
from xlsx_make import create_sanitized_xlsx
from xml_functions import basic_xml_file, dossier_entry, file_entry, volume_entry


def create_xml_individual_files(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    dossiers: dict[str, etree._Element],
    vol_entry: etree._Element,
) -> None:
    """Based on a sheet creates .xml entries for every file found"""
    for file in sheet.iter_rows():
        if file[0].value is not None and not file[0].value.endswith("_0"):
            f_page, f_title, f_place, f_date = parse_file(file)

            # Check if file belongs to a dossier
            if mat := re.search(r"ms.+?_(.+?)_.+?", file[0].value):
                file_entry(dossiers[mat.groups()[0]], f_page, f_title, f_place, f_date)
            else:
                file_entry(vol_entry, f_page, f_title, f_place, f_date)


def create_xml_dossier(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    v_num: str,
    c01: etree._Element,
) -> None:
    """Creates necessary dossier entries at the c02 level"""
    dossiers: dict[str, etree._Element] = {}

    # Find dossiers
    for cell in sheet["A"]:
        if cell.value is not None and (
            mat := re.search(r"ms.+?_(.+?)_.+?", cell.value)
        ):
            if mat.groups()[0] not in dossiers.keys():
                d_title, d_date = parse_dossier(sheet, mat.groups()[0], v_num, cell)

                # Create entry
                dossiers[mat.groups()[0]] = dossier_entry(
                    c01,
                    v_num,
                    mat.groups()[0],
                    d_title,
                    d_date,
                )

    # Create dossier if no dossiers were found?
    # TODO: Do we want to create a dossier in these cases?
    if dossiers == {}:
        warn(f"V{v_num} does not have any dossiers!")

    # create_xml_individual_files(sheet, dossiers, c01)


def create_xml_volume(
    filename: str,
    archdesc: etree._Element,
) -> None:
    """Adds a volume to an 'archdesc' element"""
    workbook = load_workbook(filename)
    first_sheet = workbook[workbook.sheetnames[0]]

    # Create volume entry at c01 level
    volume_data = parse_volume(first_sheet[1])
    c01 = volume_entry(archdesc, volume_data)

    create_xml_dossier(first_sheet, volume_data.num, c01)

    print(f"""Finished writing volume {volume_data.num}\n""")


def create_xml_file(dir_name: str) -> None:
    """Creates and writes an xml file based on directory of volumes"""
    print("Starting to create XML file!")
    root, archdesc_dsc = basic_xml_file()

    # Iterate files sorted by volume number
    for file in sorted(
        os.listdir(dir_name),
        key=lambda name: int(
            name.replace("Paesi Bassi VOLUME ", "").replace("_it_IT.xlsx", "")
        ),
    ):
        if not file.count("~$") and file.startswith("Paesi"):
            create_xml_volume(f"{dir_name}/{file}", archdesc_dsc)

    tree = etree.ElementTree(root)

    # Check if outputs directory exists and then write file
    os.makedirs(os.path.join(os.getcwd(), r"outputs"), exist_ok=True)
    tree.write(  # type: ignore # Stub doesn't recognize doctype parameter is valid
        "outputs/Legation_Archive.xml",
        pretty_print=True,
        xml_declaration=True,
        encoding="UTF-8",
        doctype="""<!DOCTYPE ead PUBLIC "+// http://ead3.archivists.org/schema/ //DTD ead3 (Encoded Archival Description (EAD) Version 3)//EN" "ead3.dtd">""",  # pylint: disable=line-too-long
    )

    print("Printed file to outputs/Legation_Archive.xml")
    print("Writing XML complete!")


if __name__ == "__main__":
    create_sanitized_xlsx("inputs/VolumesExcel/it_IT")
    create_xml_file("outputs/VolumesExcelSanitized/it_IT")