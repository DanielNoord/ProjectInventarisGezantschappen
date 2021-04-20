import os
import re

from openpyxl import load_workbook

from functions.data_create_name_string import name_string


def fill_in_xlsx(directory_name, file_name, individuals, translations, localization):
    """Fills an individual volume .xlsx file

    Args:
        directory_name (str): Name of input directory
        file_name (str): Name of input file
        individuals (dict): Dictionary of all individuals based on inputs/Individuals.json
        translations (list): Dictionaries of function, title and places translations
        localization (str): Localization abbreviation ("nl_NL", "it_IT", "en_GB")
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        if row[1].value is not None:
            date = [row[2].value, row[3].value, row[4].value]
            line = re.split(r"( |\.|,|\(|\))", row[1].value)
            for index, word in enumerate(line):
                if word.startswith("$"):
                    line[index] = name_string(individuals[word], date, translations, localization)
            line = "".join(line)

            # Clean up string
            if line[-1] in [" ", ",", "."]:  # Remove final characters
                line = line[:-1]
            line = line[0].upper() + line[1:]
            line = line.replace("( ", "(").replace(" )", ")")
            if line == " ":
                line = ""
            row[1].value = line

    new_directory = (
        directory_name.replace("VolumesExcelSanitized/", "VolumesExcelFinal/")
        .replace("VolumesExcelTranslated/", "VolumesExcelFinal/")
        .replace("inputs", "outputs")
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
    print(f"File written to {new_directory}/{file_name}")
