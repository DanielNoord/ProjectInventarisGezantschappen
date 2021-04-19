import os

from openpyxl import load_workbook


def sanitize_xlsx(directory_name, file_name):
    """Sanitize an .xlsx file

    Args:
        directory_name (str): Directory of .xlsx file
        file_name (str): Name of .xlsx file
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        for index, cell in enumerate(row):
            if cell.value is not None:
                line = cell.value

                # Clean up string
                if line == " ":
                    line = ""
                elif isinstance(line, str):
                    if line[-1] in [" ", ",", "."]:  # Remove final characters
                        line = line[:-1]
                    if not line.startswith("ms"):
                        line = line[0].upper() + line[1:]
                    line = line.replace("( ", "(").replace(" )", ")")
                row[index].value = line

    new_directory = directory_name.replace("inputs", "outputs").replace(
        "VolumesExcel/", "VolumesExcelSanitized/"
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name}")
    print(f"File written to {new_directory}/{file_name}")
