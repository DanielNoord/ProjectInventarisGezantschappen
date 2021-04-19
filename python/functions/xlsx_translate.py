import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Font


def translate_xlsx(
    directory_name, file_name, localization, translations, used_translations
):
    """Translate .xlsx file

    Args:
        directory_name (str): Directory of .xlsx file
        file_name (str): Name of .xlsx file
        localization (str): Localization string of target language
        translations (dict): Translations of common document titles
        used_translations (set): Used translations
    """
    workbook = load_workbook(f"{directory_name}/{file_name}")
    for row in workbook[workbook.sheetnames[0]].iter_rows():
        if row[1].value is not None:
            line = row[1].value

            # Check if line matches one of the patterns that are automatically translated
            for pattern, trans in translations.items():
                if pattern.match(line):
                    line = re.sub(pattern, trans[localization], line)
                    used_translations.add(pattern)
                    break
            else:
                row[1].font = Font(color="00008b")

            # Clean up string
            if line[-1] in [" ", ",", "."]:  # Remove final characters
                line = line[:-1]
            line = line[0].upper() + line[1:]
            line = line.replace("( ", "(").replace(" )", ")")
            row[1].value = line

    new_directory = (
        directory_name.replace("inputs", "outputs")
        .replace("VolumesExcelSanitized/", "VolumesExcelTranslated/")
        .replace("it_IT", f"{localization}")
    )
    os.makedirs(
        os.path.join(os.getcwd(), new_directory),
        exist_ok=True,
    )
    workbook.save(f"{new_directory}/{file_name.replace('it_IT', localization)}")
    print(f"File written to {new_directory}/{file_name.replace('it_IT', localization)}")
