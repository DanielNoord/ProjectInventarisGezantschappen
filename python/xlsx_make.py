#!/usr/bin/env python3

import json
import os
import time
from re import Pattern
from typing import Literal

from data_parsing import initialize_database_for_xml
from openpyxl import Workbook, load_workbook
from xlsx_functions import (
    add_identifier_columns,
    fill_in_xlsx,
    sanitize_xlsx,
    translate_xlsx,
)


def create_filled_xlsx(
    directory_name: str, localization: Literal["it_IT", "nl_NL", "en_GB"]
) -> None:
    """Creates the .xlsx files of all volumes in the input directory.

    Args:
        directory_name: Name of the directory with the input .xlsx files
        localization: Localization abbreviation
    """
    with open("inputs/Individuals.json", encoding="utf-8") as data_file:
        individuals_data = json.load(data_file)
    del individuals_data["$schema"]
    database = initialize_database_for_xml()

    directory = os.fsencode(directory_name)
    for file in os.listdir(directory):
        if not str(file).count("~$") and str(file).startswith("b'Paesi"):
            filename = os.fsdecode(file)
            fill_in_xlsx(
                directory_name,
                filename,
                individuals_data,
                database,
                localization,
            )


def create_xlsx_controle(directory_names: list[str]) -> None:
    """Creates a combination of the various translated .xlsx files into one "control file".

    Args:
        directory_names: Directories of the file to be merged
    """
    directory = os.fsencode(directory_names[0])
    for file in os.listdir(directory):
        if not str(file).count("~$") and str(file).startswith("b'Paesi"):
            new_file = Workbook()
            new_sheet = new_file.active
            filename = os.fsdecode(file)
            rows_original = list(load_workbook(f"{directory_names[0]}/{filename}").active.rows)
            rows_it = list(
                load_workbook(
                    f"{directory_names[1]}/{filename.replace('.xlsx', '')}_it_IT.xlsx"
                ).active.rows
            )
            rows_en = list(
                load_workbook(
                    f"{directory_names[2]}/{filename.replace('.xlsx', '')}_en_GB.xlsx"
                ).active.rows
            )
            rows_nl = list(
                load_workbook(
                    f"{directory_names[3]}/{filename.replace('.xlsx', '')}_nl_NL.xlsx"
                ).active.rows
            )
            for index, row in enumerate(rows_original):
                if row[0].value != rows_it[index][0].value:
                    raise ValueError(
                        f"Difference in document number found for {row[0].value} in original"
                    )
                new_sheet.append(tuple(i.value for i in row))
                new_sheet.append(tuple(i.value for i in rows_it[index]))
                new_sheet.append(tuple(i.value for i in rows_en[index]))
                new_sheet.append(tuple(i.value for i in rows_nl[index]))
                new_sheet.append([])
            os.makedirs(
                os.path.join(os.getcwd(), "outputs/VolumesExcelControl"),
                exist_ok=True,
            )
            new_filename = f"Control_{filename}"
            new_file.save(f"outputs/VolumesExcelControl/{new_filename}")
            print(f"Wrote file to outputs/VolumesExcelControl/{new_filename}")


def create_translated_xlsx(directory_name: str, localization: Literal["nl_NL", "en_GB"]) -> None:
    """Creates the .xlsx files with a number of easy translation already done.

    Args:
        directory_name: Name of the directory with the input .xlsx files
        localization: Localization abbreviation
    """
    database = initialize_database_for_xml()
    used_translations: set[Pattern[str]] = set()

    directory = os.fsencode(directory_name)
    for file in sorted(os.listdir(directory)):
        if not str(file).count("~$") and str(file).startswith("b'Paesi"):
            filename = os.fsdecode(file)
            translate_xlsx(
                directory_name,
                filename,
                localization,
                database,
                used_translations,
            )
    unused_trans = [i for i in database.document_titles.keys() if i not in used_translations]
    print("Done!\nFound the following unused translations:\n", unused_trans)


def create_sanitized_xlsx(directory_name: str) -> None:
    """Creates the .xlsx files with a number of easy sanitizations.

    Args:
        directory_name: Name of the directory with the input .xlsx files
    """

    directory_path = os.path.realpath(directory_name)
    for file in sorted(os.listdir(directory_path)):
        if not str(file).count("~$") and str(file).startswith("Paesi"):
            filename = os.fsdecode(file)
            sanitize_xlsx(directory_name, filename)


def create_xlsx_with_identifier_columns(directory_name: str) -> None:
    """Create .xlsx files while adding columns for identifiers."""
    database = initialize_database_for_xml()
    # Populate surnames dict with common ways to write a name.
    surnames = {
        "Baron de Perglass": ["$Welden"],
        "Baron de Tuyll de Serooskerken chambellan de Sa": ["$TuyllGMK"],
        "Card. Soglia": ["$Soglia"],
        "Duc de Rignano": ["$Rignano"],
        "Francesco e Rosa Maiai": ["$MadiaiR", "$MadiaiR"],
        "le Prince Castelcicala": ["$RuffoP"],
        "Lord Aberdeen": ["$Aberdeen"],
        "Marechal de Welden": ["$Welden"],
        "Papa IX": ["$PioIX"],
        "Papa Neuf": ["$PioIX"],
        "Papae IX": ["$PioIX"],
        "Papae XVI": ["$GregorioXVI"],
        "Pie IX": ["$PioIX"],
        "Pie Neuf": ["$PioIX"],
        "Pio Nono": ["$PioIX"],
        "Pius PP. IX": ["$PioIX"],
        "Prince de Capoue": ["$BourbonKF"],
        "Terenzio Mamiani": ["$Mamiani"],
        "Vittorio Emanuele II": ["$SavoiaVE"],
    }
    for identifier, data in database.individuals.items():
        if data["surname"] in surnames:
            surnames[data["surname"]].append(identifier)
        else:
            surnames[data["surname"]] = [identifier]
    del surnames["Rap"]  # Matches with all 'Rapporto ...' titles
    directory_path = os.path.realpath(directory_name)
    for file in sorted(os.listdir(directory_path)):
        if not str(file).count("~$") and str(file).startswith("Paesi"):
            filename = os.fsdecode(file)
            add_identifier_columns(directory_name, filename, surnames)


def do_full_loop() -> None:
    """Completes the full process of input files till seperate translations and control file."""
    print("STARTING CREATION OF .XLSX DOCUMENTS\n")
    start_time = time.time()
    create_sanitized_xlsx("inputs/VolumesExcel/it_IT")
    create_translated_xlsx("outputs/VolumesExcelSanitized/it_IT", "en_GB")
    create_translated_xlsx("outputs/VolumesExcelSanitized/it_IT", "nl_NL")
    create_filled_xlsx("outputs/VolumesExcelTranslated/en_GB", "en_GB")
    create_filled_xlsx("outputs/VolumesExcelSanitized/it_IT", "it_IT")
    create_filled_xlsx("outputs/VolumesExcelTranslated/nl_NL", "nl_NL")
    create_xlsx_controle(
        [
            "inputs/VolumesExcel/original",
            "outputs/VolumesExcelFinal/it_IT",
            "outputs/VolumesExcelFinal/en_GB",
            "outputs/VolumesExcelFinal/nl_NL",
        ]
    )
    print(
        "\n\nFINISHED CREATING ALL .XLSX DOCUMENTS\n",
        f"Process took {time.time() - start_time:.03f} seconds",
    )


if __name__ == "__main__":
    create_xlsx_with_identifier_columns("inputs/VolumesExcel/it_IT")
    # create_sanitized_xlsx("inputs/VolumesExcel/it_IT")
    # create_translated_xlsx("outputs/VolumesExcelSanitized/it_IT", "en_GB")
    # create_translated_xlsx("outputs/VolumesExcelSanitized/it_IT", "nl_NL")
    # create_filled_xlsx("outputs/VolumesExcelTranslated/en_GB", "en_GB")
    # create_filled_xlsx("outputs/VolumesExcelSanitized/it_IT", "it_IT")
    # create_filled_xlsx("outputs/VolumesExcelTranslated/nl_NL", "nl_NL")
    # create_xlsx_controle(
    #     [
    #         "inputs/VolumesExcel/original",
    #         "outputs/VolumesExcelFinal/it_IT",
    #         "outputs/VolumesExcelFinal/en_GB",
    #         "outputs/VolumesExcelFinal/nl_NL",
    #     ]
    # )
    # do_full_loop()
